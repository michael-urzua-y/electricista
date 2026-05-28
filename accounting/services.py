"""
Servicios de contabilidad para generar libros de compras, ventas y resúmenes mensuales.
"""
import re
from decimal import Decimal
from typing import List, Dict, Optional
from io import BytesIO

from django.db.models import Sum, Q
from django.contrib.auth import get_user_model

from invoices.models import Invoice
from quotes.models import Quote
from monaysolutions.config import IVA_RATE, TWO_PLACES

User = get_user_model()


def _extract_rut_from_ocr(ocr_text: str) -> str:
    """
    Extrae el RUT del PROVEEDOR desde el texto OCR.
    Busca patrones como 'R.U.T. 81.515.100-3' que aparecen en el encabezado
    de la factura (razón social del emisor), no el RUT del cliente.
    Solo retorna si aparece junto a palabras clave del emisor.
    """
    if not ocr_text:
        return ''
    # Buscar RUT que aparece junto a "R.U.T." (formato típico del emisor en facturas chilenas)
    # Excluir el RUT que aparece después de "SEÑOR(ES)", "RUT:" (que es del cliente)
    lines = ocr_text.split('\n')
    for i, line in enumerate(lines):
        line_clean = line.strip()
        # Patrón del emisor: "R.U.T. XXXXXXXX-X" o "R.U.T XXXXXXXX-X"
        if re.match(r'R\.U\.T\.?\s*\d', line_clean, re.IGNORECASE):
            match = re.search(r'(\d{1,2}[\.\d]*\d{3}[\.\d]*\d{3}-[\dkK])', line_clean)
            if match:
                return match.group(1).strip()
        # También buscar en línea siguiente si la línea actual es solo "R.U.T."
        if re.match(r'^R\.U\.T\.?\s*$', line_clean, re.IGNORECASE) and i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            match = re.search(r'(\d{1,2}[\.\d]*\d{3}[\.\d]*\d{3}-[\dkK])', next_line)
            if match:
                return match.group(1).strip()
    return ''


def _calc_neto_iva(invoice: Invoice):
    """
    Calcula neto e IVA desde los campos de la factura.
    Si subtotal_amount y tax_amount están disponibles, los usa directamente.
    Si no, calcula desde total_amount usando el IVA configurado.
    """
    total = invoice.total_amount or Decimal('0')

    if invoice.subtotal_amount and invoice.tax_amount:
        return invoice.subtotal_amount, invoice.tax_amount

    if invoice.subtotal_amount:
        neto = invoice.subtotal_amount
        iva = total - neto
        return neto, iva

    if invoice.tax_amount:
        iva = invoice.tax_amount
        neto = total - iva
        return neto, iva

    # Calcular desde total: total = neto * (1 + IVA)
    neto = (total / (1 + IVA_RATE)).quantize(TWO_PLACES)
    iva = total - neto
    return neto, iva


def get_libro_compras(user, year: int, month: int) -> List[Dict]:
    """
    Genera el libro de compras mensual con todas las facturas de proveedores del período.
    
    Args:
        user: Usuario autenticado
        year: Año del período
        month: Mes del período (1-12)
    
    Returns:
        Lista de diccionarios con: rut_proveedor, razon_social, folio, fecha, neto, iva, total
    """
    # Filtrar facturas completadas del período
    invoices = Invoice.objects.filter(
        user=user,
        status='completed',
        issue_date__year=year,
        issue_date__month=month
    ).select_related('provider').order_by('issue_date', 'invoice_number')
    
    libro = []
    for invoice in invoices:
        # RUT: primero del proveedor, luego del OCR
        rut_proveedor = getattr(invoice.provider, 'rut', '') or ''
        if not rut_proveedor and invoice.ocr_text:
            rut_proveedor = _extract_rut_from_ocr(invoice.ocr_text)
        sin_rut = not bool(rut_proveedor)
        if not rut_proveedor:
            rut_proveedor = 'Sin RUT'

        razon_social = invoice.provider.name if invoice.provider else 'Sin proveedor'

        # Neto e IVA calculados correctamente
        neto, iva = _calc_neto_iva(invoice)
        total = invoice.total_amount or Decimal('0')

        libro.append({
            'rut_proveedor': rut_proveedor,
            'razon_social': razon_social,
            'folio': invoice.invoice_number or f'INV-{invoice.id}',
            'fecha': invoice.issue_date,
            'neto': neto,
            'iva': iva,
            'total': total,
            'sin_rut': sin_rut,
        })
    
    return libro


def get_libro_ventas(user, year: int, month: int) -> List[Dict]:
    """
    Genera el libro de ventas mensual con todas las cotizaciones aprobadas del período.
    
    Args:
        user: Usuario autenticado
        year: Año del período
        month: Mes del período (1-12)
    
    Returns:
        Lista de diccionarios con: rut_cliente, nombre_cliente, folio, fecha, neto, iva, total
    """
    # Filtrar cotizaciones aprobadas del período
    # Usamos status_updated_at para la fecha de aprobación, con fallback a created_at
    quotes = Quote.objects.filter(
        user=user,
        status='approved'
    ).filter(
        Q(status_updated_at__year=year, status_updated_at__month=month) |
        Q(status_updated_at__isnull=True, created_at__year=year, created_at__month=month)
    ).select_related('client').order_by('status_updated_at', 'quote_number')
    
    libro = []
    for quote in quotes:
        rut_cliente = quote.client_rut or (quote.client.rut if quote.client else '')
        nombre_cliente = quote.client_name or (quote.client.name if quote.client else 'Sin cliente')
        fecha_aprobacion = quote.status_updated_at.date() if quote.status_updated_at else quote.created_at.date()

        # Neto e IVA desde los campos de la cotización
        neto = quote.subtotal or Decimal('0')
        iva = quote.tax_amount or Decimal('0')
        total = quote.total_amount or Decimal('0')

        # Si neto es 0 pero total no, calcular
        if not neto and total:
            neto = (total / (1 + IVA_RATE)).quantize(Decimal('0.01'))
            iva = total - neto

        sin_rut = not bool(rut_cliente)
        if not rut_cliente:
            rut_cliente = 'Sin RUT'

        libro.append({
            'rut_cliente': rut_cliente,
            'nombre_cliente': nombre_cliente,
            'folio': quote.quote_number,
            'fecha': fecha_aprobacion,
            'neto': neto,
            'iva': iva,
            'total': total,
            'sin_rut': sin_rut,
        })
    
    return libro


def get_resumen_mensual(user, year: int, month: int) -> Dict:
    """
    Genera el resumen contable mensual consolidado de compras, ventas e IVA.
    
    Args:
        user: Usuario autenticado
        year: Año del período
        month: Mes del período (1-12)
    
    Returns:
        Diccionario con totales de compras, ventas, IVA y comparación con mes anterior
    """
    # Obtener libros del mes actual
    libro_compras = get_libro_compras(user, year, month)
    libro_ventas = get_libro_ventas(user, year, month)
    
    # Calcular totales del mes actual
    total_compras_netas = sum(item['neto'] for item in libro_compras)
    total_iva_soportado = sum(item['iva'] for item in libro_compras)
    total_compras_brutas = sum(item['total'] for item in libro_compras)
    
    total_ventas_netas = sum(item['neto'] for item in libro_ventas)
    total_iva_debito = sum(item['iva'] for item in libro_ventas)
    total_ventas_brutas = sum(item['total'] for item in libro_ventas)
    
    # Resultado IVA (débito - soportado)
    resultado_iva = total_iva_debito - total_iva_soportado
    
    # Calcular mes anterior
    mes_anterior = month - 1 if month > 1 else 12
    año_anterior = year if month > 1 else year - 1
    
    # Obtener totales del mes anterior
    libro_compras_anterior = get_libro_compras(user, año_anterior, mes_anterior)
    libro_ventas_anterior = get_libro_ventas(user, año_anterior, mes_anterior)
    
    total_compras_netas_anterior = sum(item['neto'] for item in libro_compras_anterior)
    total_ventas_netas_anterior = sum(item['neto'] for item in libro_ventas_anterior)
    total_iva_soportado_anterior = sum(item['iva'] for item in libro_compras_anterior)
    total_iva_debito_anterior = sum(item['iva'] for item in libro_ventas_anterior)
    resultado_iva_anterior = total_iva_debito_anterior - total_iva_soportado_anterior
    
    # Calcular variaciones porcentuales
    def calcular_variacion(actual, anterior):
        if anterior == 0:
            return None  # Evitar división por cero
        return ((actual - anterior) / anterior) * 100
    
    return {
        # Mes actual
        'total_compras_netas': total_compras_netas,
        'total_ventas_netas': total_ventas_netas,
        'iva_soportado': total_iva_soportado,
        'iva_debito': total_iva_debito,
        'resultado_iva': resultado_iva,
        'total_compras_brutas': total_compras_brutas,
        'total_ventas_brutas': total_ventas_brutas,
        
        # Mes anterior
        'total_compras_netas_anterior': total_compras_netas_anterior,
        'total_ventas_netas_anterior': total_ventas_netas_anterior,
        'iva_soportado_anterior': total_iva_soportado_anterior,
        'iva_debito_anterior': total_iva_debito_anterior,
        'resultado_iva_anterior': resultado_iva_anterior,
        
        # Variaciones
        'variacion_compras': calcular_variacion(total_compras_netas, total_compras_netas_anterior),
        'variacion_ventas': calcular_variacion(total_ventas_netas, total_ventas_netas_anterior),
        'variacion_iva': calcular_variacion(resultado_iva, resultado_iva_anterior),
        
        # Indicadores
        'iva_a_pagar': resultado_iva > 0,
        'monto_iva': abs(resultado_iva),
    }


def export_to_excel(data: List[Dict], columns: List[str], sheet_name: str = 'Hoja1') -> bytes:
    """
    Genera un archivo Excel (.xlsx) a partir de una lista de diccionarios.
    
    Args:
        data: Lista de diccionarios con los datos a exportar
        columns: Lista de nombres de columnas en el orden deseado
        sheet_name: Nombre de la hoja de Excel
    
    Returns:
        Bytes del archivo Excel generado
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecute: pip install openpyxl")
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Mapeo de nombres de columnas a títulos legibles
    column_titles = {
        'rut_proveedor': 'RUT Proveedor',
        'razon_social': 'Razón Social',
        'rut_cliente': 'RUT Cliente',
        'nombre_cliente': 'Nombre Cliente',
        'folio': 'Folio',
        'fecha': 'Fecha',
        'neto': 'Neto',
        'iva': 'IVA',
        'total': 'Total',
    }
    
    # Escribir encabezados
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column_titles.get(col_name, col_name.replace('_', ' ').title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir datos
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col_name, '')
            
            # Formatear valores
            if col_name == 'fecha' and value:
                cell.value = value.strftime('%d/%m/%Y') if hasattr(value, 'strftime') else str(value)
            elif col_name in ['neto', 'iva', 'total']:
                cell.value = float(value) if value else 0
                cell.number_format = '#,##0'
            else:
                cell.value = str(value) if value else ''
            
            # Resaltar filas sin RUT
            if row_data.get('sin_rut', False) and col_name in ['rut_proveedor', 'rut_cliente']:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                cell.font = Font(color='9C0006')
    
    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = len(column_titles.get(col_name, col_name))
        
        # Calcular ancho basado en contenido
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Agregar fila de totales si hay datos numéricos
    if data and any(col in columns for col in ['neto', 'iva', 'total']):
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1).value = 'TOTAL'
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name in ['neto', 'iva', 'total']:
                cell = ws.cell(row=total_row, column=col_idx)
                total = sum(row_data.get(col_name, 0) for row_data in data)
                cell.value = float(total)
                cell.number_format = '#,##0'
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
