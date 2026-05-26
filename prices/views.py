"""
Vistas para el módulo de precios.
"""
import io
from django.db.models import Count
from django.db import transaction
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook

from .models import PriceItem, PriceSubItem
from .serializers import (
    PriceItemSerializer,
    PriceItemListSerializer,
    PriceSubItemSerializer,
    PriceSubItemSearchSerializer,
)


class PriceItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Ítems de precio.

    list:    GET    /api/prices/items/
    create:  POST   /api/prices/items/
    retrieve: GET   /api/prices/items/{id}/
    update:  PUT    /api/prices/items/{id}/
    destroy: DELETE /api/prices/items/{id}/
    subitems: GET/POST /api/prices/items/{id}/subitems/
    """
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return (
            PriceItem.objects
            .filter(user=self.request.user)
            .annotate(subitems_count=Count('subitems'))
            .order_by('order_number')
        )

    def get_serializer_class(self):
        if self.action == 'list':
            return PriceItemListSerializer
        return PriceItemSerializer

    def perform_create(self, serializer):
        # Asignar el siguiente número de orden disponible
        last_order = (
            PriceItem.objects
            .filter(user=self.request.user)
            .order_by('-order_number')
            .values_list('order_number', flat=True)
            .first()
        ) or 0
        serializer.save(user=self.request.user, order_number=last_order + 1)

    def perform_destroy(self, instance):
        # Elimina el ítem y sus sub-ítems (CASCADE)
        instance.delete()

    # ------------------------------------------------------------------
    # Action: subitems — listar y crear sub-ítems de un ítem
    # ------------------------------------------------------------------
    @action(detail=True, methods=['get', 'post'], url_path='subitems')
    def subitems(self, request, pk=None):
        item = self.get_object()

        if request.method == 'GET':
            subitems = item.subitems.all().order_by('sub_number')
            serializer = PriceSubItemSerializer(subitems, many=True)
            return Response(serializer.data)

        # POST — crear sub-ítem
        serializer = PriceSubItemSerializer(data=request.data)
        if serializer.is_valid():
            # Asignar siguiente número secuencial
            last_sub = (
                item.subitems
                .order_by('-sub_number')
                .values_list('sub_number', flat=True)
                .first()
            ) or 0
            serializer.save(item=item, sub_number=last_sub + 1)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # ------------------------------------------------------------------
    # Action: download_template — descargar plantilla Excel para carga masiva
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla de Precios"

        # Encabezados
        ws['A1'] = "Ítem Orden"
        ws['B1'] = "Ítem Nombre"
        ws['C1'] = "Sub-ítem Número"
        ws['D1'] = "Sub-ítem Descripción"
        ws['E1'] = "Valor Neto"

        # Ejemplo: Ítem 1 con 2 sub-ítems
        ws['A2'] = 1
        ws['B2'] = "PUNTO DE RED"
        ws['C2'] = 1
        ws['D2'] = "Instalación de Punto de red Cat 6 en altura, Incluye certificación."
        ws['E2'] = 10000

        ws['A3'] = 1
        ws['B3'] = "PUNTO DE RED"
        ws['C3'] = 2
        ws['D3'] = "Instalación de Punto de red Cat 6A AP en altura, Incluye certificación."
        ws['E3'] = 5000

        # Ejemplo: Ítem 2 con 2 sub-ítems
        ws['A4'] = 2
        ws['B4'] = "FIBRA ÓPTICA"
        ws['C4'] = 1
        ws['D4'] = "Tendido y provisión de Fibra óptica multimodo OM3 6F."
        ws['E4'] = 5000

        ws['A5'] = 2
        ws['B5'] = "FIBRA ÓPTICA"
        ws['C5'] = 2
        ws['D5'] = "Fusión de fibra óptica en modalidad termo-fusión más certificación, entrega de documentación PDF a cliente."
        ws['E5'] = 5400

        # Ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 70
        ws.column_dimensions['E'].width = 15

        # Preparar respuesta como HttpResponse (no DRF Response) para binarios
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_precios.xlsx"'
        return response

    # ------------------------------------------------------------------
    # Action: upload_excel — procesar archivo Excel para crear/actualizar precios
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='upload-excel')
    def upload_excel(self, request):
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó ningún archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        excel_file = request.FILES['file']
        if not excel_file.name.endswith('.xlsx'):
            return Response(
                {'error': 'El archivo debe ser formato .xlsx'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            item_cache = {}
            created_items = 0
            created_subitems = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    item_order, item_name, sub_number, description, net_value = row

                    if item_order is None or item_name is None:
                        errors.append(f"Fila {row_idx}: Ítem orden y nombre son requeridos")
                        continue

                    item_key = (int(item_order), str(item_name).strip())
                    if item_key not in item_cache:
                        item, created = PriceItem.objects.get_or_create(
                            user=request.user,
                            order_number=item_order,
                            defaults={'name': item_name}
                        )
                        if not created:
                            if item.name != item_name:
                                item.name = item_name
                                item.save(update_fields=['name'])
                        else:
                            created_items += 1
                        item_cache[item_key] = item.id
                    else:
                        item = PriceItem.objects.get(id=item_cache[item_key])

                    if sub_number is not None and description is not None and net_value is not None:
                        sub_item, sub_created = PriceSubItem.objects.get_or_create(
                            item=item,
                            sub_number=int(sub_number),
                            defaults={
                                'description': str(description).strip(),
                                'net_value': net_value
                            }
                        )
                        if not sub_created:
                            updated = False
                            if sub_item.description != str(description).strip():
                                sub_item.description = str(description).strip()
                                updated = True
                            if sub_item.net_value != net_value:
                                sub_item.net_value = net_value
                                updated = True
                            if updated:
                                sub_item.save(update_fields=['description', 'net_value'])
                        if sub_created:
                            created_subitems += 1
                    elif sub_number is not None or description is not None or net_value is not None:
                        errors.append(f"Fila {row_idx}: Para sub-ítem se requieren número, descripción y valor neto")

                except Exception as e:
                    errors.append(f"Fila {row_idx}: Error al procesar - {str(e)}")

            if errors and not (created_items or created_subitems):
                return Response(
                    {'error': 'Errores en el archivo', 'details': errors[:10]},
                    status=status.HTTP_400_BAD_REQUEST
                )

            message = []
            if created_items:
                message.append(f"{created_items} categoría(s) creada(s)")
            if created_subitems:
                message.append(f"{created_subitems} sub-ítem(es) creado(s)")

            return Response({
                'message': ', '.join(message) if message else 'No se encontraron datos nuevos para crear',
                'errors': errors[:10] if errors else []
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PriceSubItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión directa de Sub-Ítems (update, delete).

    list:    GET    /api/prices/subitems/       - todos los sub-items del usuario
    retrieve: GET  /api/prices/subitems/{id}/
    update:  PUT   /api/prices/subitems/{id}/
    destroy: DELETE /api/prices/subitems/{id}/

    Búsqueda: GET /api/prices/subitems/?search=<texto>
    - Si search tiene < 2 caracteres, retorna lista vacía
    - Filtra por description__icontains y limita a 20 resultados
    """
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        search = self.request.query_params.get('search', '').strip()
        if search:
            return PriceSubItemSearchSerializer
        return PriceSubItemSerializer

    def get_queryset(self):
        search = self.request.query_params.get('search', '').strip()
        if search:
            if len(search) < 2:
                return PriceSubItem.objects.none()
            return (
                PriceSubItem.objects
                .filter(item__user=self.request.user, description__icontains=search)
                .select_related('item')
            )[:20]
        return (
            PriceSubItem.objects
            .filter(item__user=self.request.user)
            .select_related('item')
            .order_by('item__order_number', 'sub_number')
        )

    @property
    def paginator(self):
        """Disable pagination when search param is present."""
        search = self.request.query_params.get('search', '').strip()
        if search:
            return None
        return super().paginator

    def perform_update(self, serializer):
        serializer.save()
